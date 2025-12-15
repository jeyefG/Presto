# -*- coding: utf-8 -*-
"""
Created on Mon Oct 17 10:57:47 2022
Obras:
    Proyectos Ok: República, FDM, Dragpharma, Conosur, CCU Serie 300, Viña Carmen (buin)
    Chequear que no existan cuadrillas de MO, y que sucontratos sean de tipo "Otros" y no partidas
@author: jfcog
"""

from openpyxl import load_workbook
import pandas as pd
import copy
import os
import tkinter as tk
from tkinter import filedialog


root = tk.Tk()
root.withdraw()

# Crea una ventana temporal para forzar el enfoque
temp_window = tk.Toplevel()
temp_window.withdraw()  # Oculta la ventana temporal
temp_window.focus_force()  # Fuerza el enfoque en la ventana temporal

file = filedialog.askopenfilename()
path = os.path.dirname(file)
wb = load_workbook(filename = path + '/Libro1.xlsx')
sheet_names = wb.sheetnames
name = sheet_names[0]
sheet_ranges = wb[name]
df = pd.DataFrame(sheet_ranges.values)
root.destroy()

num_cap = 6
num_part = 4
rec_type = ["Otros","Material","Mano de obra","Maquinaria"]

def sum_to_list(rango):
    
    nuevas_filas = []
    aux1 = rango[rango.index('(')+1:rango.index(')')]
    aux2 = aux1[aux1.index('G')+1:aux1.index(':')]
    aux3 = aux1[aux1.index(':')+2:]
    inicio = int(aux2)
    fin = int(aux3)+1
    for i in range(inicio,fin):
        celda = 'G' + str(i)
        nuevas_filas.append(celda)
    #print(nuevas_filas)
    return nuevas_filas
#revisa si el elemento de una celda sólo nos cambia de posición hasta que llegue a una sum or Round
#devuelve las formulas de la celda que no son solo un cambio de posición
def position_change_for(elementn_for):
    while elementn_for[0:1] == "E" or elementn_for[0:1] == "F" or elementn_for[0:1] == "G":
        
        new_elementn_col = 0 
        new_elementn_row = 0
        if elementn_for[0:1] == "E":
            new_elementn_col = 4
            new_elementn_row = int(elementn_for[1:]) -1
            elementn = df.loc[new_elementn_row][new_elementn_col]  
            return elementn
        if elementn_for[0:1] == "F":
            new_elementn_col = 5
            new_elementn_row = int(elementn_for[1:]) -1
        if elementn_for[0:1] == "G":
            new_elementn_col = 6        #formulas están en columna F, en la G sólo redondea
            new_elementn_row = int(elementn_for[1:]) -1
            
        elementn = df.loc[new_elementn_row][new_elementn_col]  

        if isinstance(elementn,float) or isinstance(elementn,int):
            x = 0
        else:
            elementn_for = list(elementn.split("+"))
            aux9 = 0    
            for elemente in elementn_for:
                if elemente[0] == "=":
                    elementn_for[aux9] = elemente[1:]
                
                aux9 =+ 1
            #print(elementn_for)
            break

    
    return elementn_for
def check_sum_elements(element):
    
    formulascs = []
    while '+' in element:
        
        string_aux = element[0:element.index('+')]
        #print(string_aux)
        
        if '=' in string_aux or '+' in string_aux:
            string_aux = string_aux[1:]
        
        formulascs.append(string_aux)
        element = element[element.index('+')+1:]
        last_for = element
    
    formulascs.append(last_for)
    return formulascs
        
#chequea si formula redondear es sobre la multiplicación de 2 números, o si es sobre más fórmulas
def check_final_round(element):
    form_cel = ""
    datos = []
    result = False
    ast = element.index('*')
    par1 = element.index('(')
    coma = element.index(',')
    cell1 = element[par1+1:ast]
    cell2 = element[ast+1:coma]
    rowcell1 = int(cell1[1:])
    rowcell2 = int(cell2[1:])
    colcell1 = 0
    colcell2 = 0
    if cell1[0:1] == 'E':
        colcell1 = 4
    if cell1[0:1] == 'F':
        colcell1 = 5
    if cell1[0:1] == 'G':
        colcell1 = 6
    
    if cell2[0:1] == 'E':
        colcell2 = 4
    if cell2[0:1] == 'F':
        colcell2 = 5
    if cell2[0:1] == 'G':
        colcell2 = 6
    
        
    cell1cont = df.loc[rowcell1-1][colcell1]
    cell2cont = df.loc[rowcell2-1][colcell2]
    #print(cell1, cell2, cell1cont, cell2cont)
    if (isinstance(cell1cont,float) or isinstance(cell1cont,int)) and (isinstance(cell2cont,float) or isinstance(cell2cont,int)):
        result = True
        datos.append(cell1cont)
        datos.append(cell2cont)
    else:
        result = False
        if (isinstance(cell1cont,float) or isinstance(cell1cont,int)):
            form_cel = cell2cont
        if (isinstance(cell2cont,float) or isinstance(cell2cont,int)):
            form_cel = cell1cont
            
        
    return (result, form_cel, datos)

#identifica las celdas del siguiente nivel, dado una celda de un nivel superior
def final_form_list(element):
    
    final_list = []
    formulas = list(element.split('+'))
    flag = False
    #chequeamos si la fórmula es un cambio de línea o no
    if len(formulas) == 1:
        while formulas[0][0:1] == 'E' or formulas[0][0:1] == 'F' or formulas[0][0:1] == 'G':
            item = formulas[0]
            formulas = position_change_for(item)
        
    for elem in formulas:
        if elem[0:3] == 'SUM':
            final_list.append(sum_to_list(elem))
            continue
        if elem[0:5] == 'ROUND':
            if check_final_round(elem)[0]:
                final_list = check_final_round(elem)[2]
                flag = True
            if not check_final_round(elem)[0]:
                final_list = check_final_round(elem)[1].split('+')
                aux = final_list[0][1:]  #sacamos signo =
                if aux[0:3] == 'SUM':
                    final_list = sum_to_list(aux)
                    continue
            continue
        
        final_list.append(elem)
        
    if not flag:
        if final_list[0][0:1] =='=':
            final_list[0] = final_list[0][1:]
        
    return final_list

def final_cant(element):
    
    formulas = list(element.split('+'))
    #chequeamos si la fórmula es un cambio de línea o no
    if len(formulas) == 1:
        while formulas[0][0:1] == 'E' or formulas[0][0:1] == 'F' or formulas[0][0:1] == 'G':
            item = formulas[0]
            formulas = position_change_for(item)
            if isinstance(formulas,int) or isinstance(formulas,float):
                break
    return formulas

def map_struct(struct, rec_um, rec_precio, ed):
    #primero revisamos cuántos niveles hay:
    levels = len(struct)
    # la estructura a llenar es fija. cant_cap + cant_part + 1 recurso+ UM recurso + cant de cada cap + cant cada partida + cant rec + precio recurso
    # el nombre, cantidad y UM del recurso se llena en cada caso. 
    niveles, nombres, cantidades = list(zip(*struct))
    niveles = list(niveles)
    nombres = list(nombres)
    cantidades = list(cantidades)
    chap_index = niveles.index("Capítulo")
    actual_chap = 0
    actual_part = 0
    for level in niveles:
        if level == "Capítulo":
            actual_chap += 1
        if level == "Partida":
            actual_part += 1
    
    total_cant = 2*num_cap + 2*num_part + 1 + 1 + 1 + 1
    
    line = []
    chap_cant = []
    part_cant = []
    #Llenamos la línea con los capítulos
    for i in range(chap_index,actual_chap):
        line.append(nombres[i])
        chap_cant.append(cantidades[i])

    #Llenamos los espacios de los niveles sin capítulos
    while len(line) < num_cap:
        line.append("")
        chap_cant.append(1)
    
    #llenamos la línea con las partidas
    try:
        part_index = niveles.index("Partida")
        for i in range(part_index, part_index + actual_part):
            line.append(nombres[i])
            part_cant.append(cantidades[i])

    except ValueError:
        pass

    #llenamos los espacios de los niveles sin partidas
    while len(line) < num_cap + num_part:
        line.append("")
        part_cant.append(1)
    
    rec_flag = 0
    for rec in rec_type:
        try:
            index = niveles.index(rec)
            line.append(nombres[index])    
            line.append(niveles[index])
            rec_flag = 1
        except ValueError:
            pass
    
    if rec_flag == 0:
        line.append("")
        line.append("")
        
    line.append(rec_um)
                    
    for element in chap_cant:
        line.append(element)
    for element in part_cant:
        line.append(element)
        
    rec_flag_cant = 0
    for rec in rec_type:
        try:
            index = niveles.index(rec)
            line.append(cantidades[index])    
            rec_flag_cant = 1
        except ValueError:
            pass
    
    if rec_flag_cant == 0:
        line.append(1)
        
    
    line.append(rec_precio)
    
    total_cant = 1
    for i in range (num_cap+num_part+2+1,2*num_cap+2*num_part+2+1+1):
        total_cant = total_cant*line[i]
        
    line.append(total_cant)
    line.append(total_cant*rec_precio)
    line.append(ed)
    line.append(priceID)
    line.append(priceIDpos)
    print(line)
    
    return line
    
    
#INICIO

    
final_map = []
header = ("Capitulo 1", "Capitulo 2 ", "Capitulo 3", "Capitulo 4", "Capitulo 5", "Capitulo 6", "Partida 1", "Partida 2", "Partida 3", "Partida 4", "Recurso", "Tipo Recurso", "UM Recurso", "Cantidad C1", "Cantidad C2", "Cantidad C3", "Cantidad C4", "Cantidad C5", "Cantidad C6", "Cantidad P1", "Cantidad P2", "Cantidad P3", "Cantidad P4", "Cantidad Recurso", "Precio Recurso","Total Cant","Presupuesto Total", "Editor")
final_map.append(header)
map_row_count = 0
priceID = 0
priceIDpos = 0

#posición del valor total del presupuesto:
#1. Los encabezados del presto están en la fila #2
#2. La fórmula del presupuesto está en la columna de Pres
#3. ültima fila es una fila vacía, el total está en la penúltima

nat_col = 1
um_col = 2
name_col = 3
cant_col = 4
pres_col = 5
colppto = 0
rowppto = 0
target_row = 686
aux = 0
for values in df.loc[2]:
    aux
    if df.loc[2][aux] == "Pres":
        colppto = aux
        break
    aux+= 1
    
#la fila del presupuesto es:

rowppto = len(df.index)-2

#separamos la formula del presupuesto en una lista, separando el string por el 
#delimitador +. La lista queda en orden inverso

presupuesto = df.loc[rowppto][colppto]
#level1_for = list(presupuesto.split("+"))

level1_for = final_form_list(presupuesto)

#listado de fórmulas donde están los totales de los items del ppto:}
#print("Level 1:")
#print(level1_for)        

#Ahora recorremos todas las celdas indicadas en items_for
#cada elemento nos dirige a la fórmula donde se construye el item respectivo

flag1 = 0
flag2 = 0
write1 = 0

#loop #1 para los items de Level1.   
for item1 in level1_for:
    
    struct = []
    priceID += 1
    priceIDpos = 1
    if map_row_count == target_row:
        x = 0
    #nombre del nivel
    flag11 = 0
    level1_nat = df.loc[int(item1[1:])-1][nat_col]
    level1_name = df.loc[int(item1[1:])-1][name_col]
    level1_cant = df.loc[int(item1[1:])-1][cant_col]
    level1_um = df.loc[int(item1[1:])-1][um_col]
    flag1 = 1
    if flag1 == 1 or flag2 == 1:
        priceID += 1
        priceIDpos = 2
    if not isinstance(level1_cant,int) and not isinstance(level1_cant,float):
        if level1_cant[0:1] == '=':
            l1cant = level1_cant[1:]
            level1_cant = final_cant(l1cant)
        struct.append([level1_nat,level1_name,level1_cant])
        priceID += 1
        priceIDpos = 3
        
    #final form nos entrega todas las formulas que componen el nivel 2 para element
    level2_for = final_form_list(item1)
    if isinstance(level2_for[0],int) or isinstance(level2_for[0],float):
        #hemos llegado a los datos finales. Hay que escribir el mapa
        struct.append([level1_nat,level1_name,level2_for[0]])
        final_map.append(map_struct(struct, level1_um, level2_for[1],"a"))
        map_row_count += 1
        #datos finales en este nivel significa que solo hay Capitulo 1 y nada mas, por lo que datos finales son un precio unitario
        priceID += 1
        priceIDpos = 4
        print(map_row_count)
        write1 = 1
        continue
    
    #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
    if len(level2_for) == 1:
        itemx= level2_for[0]
        level2_nat = df.loc[int(itemx[1:])-1][nat_col]
        level2_name = df.loc[int(itemx[1:])-1][name_col]
        level2_cant = df.loc[int(itemx[1:])-1][cant_col]
        level2_um = df.loc[int(itemx[1:])-1][um_col]
        level2_pres = df.loc[int(itemx[1:])-1][pres_col]
        level2_for = final_form_list(itemx)
        level2_name_aux = level2_name
        flag11 = 1
        #struct.append([level3_nat,level3_name,level3_cant])
    
        if isinstance(level2_for[0],int) or isinstance(level2_for[0],float):
            #hemos llegado a los datos finales. Hay que escribir el mapa     
            struct.append([level2_nat,level2_name,level2_cant])
            final_map.append(map_struct(struct, level2_um, level2_pres,"b"))
            map_row_count += 1
            print(map_row_count)
            #datos finales en este nivel significa que solo hay Capitulo 1 y nada mas, por lo que datos finales son un precio unitario
            priceID += 1
            priceIDpos = 5
            continue
        if level2_nat == "Partida":
            flag1 = 1
            if not isinstance(level2_cant,int) and not isinstance(level2_cant,float):
                if level2_cant[0:1] == '=':
                    l2cant = level2_cant[1:]
                    level2_cant = final_cant(l2cant)
            struct.append([level2_nat,level2_name,level2_cant])
        
    #revisar si el largo de la lista de nivel es = 1, ya que podría ser que sea directamente una partida
    
    #loop #2 para recorrer los capítulos del Level2
    aux2 = 0
    flag2 = 0
    flag3 = 0
    flag22 = 0
    write2 = 0
    write3 = 0
    write4 = 0
    write5 = 0

    for item2 in level2_for:
        
        if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
            
            niveles, nombres, cantidades = list(zip(*struct))
            
            if write2 == 1 and (struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida" and struct[-1][2] == 0):
                priceID += 1
                priceIDpos = 6

            if write2 == 1 and flag11 == 1:
                #se creo un nivel adicional
                struct = struct[:nombres.index(level2_name_aux)+1]
                if flag2 == 0:
                    priceID += 1
                    priceIDpos = 7
            else:
                struct = struct[:nombres.index(level1_name)+1]
                if (flag1== 0 and flag2 == 0) or flag22 == 1:
                    priceID += 1
                    priceIDpos = 8
                
                if flag3 == 1 and write3 == 1:
                    priceID += 1
                    priceIDpos = 9
                
        
        elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and (write2 == 1 or write3 == 1 or write4 == 1 or write5 == 1) :
            
            niveles, nombres, cantidades = list(zip(*struct))

            struct = struct[:nombres.index(level1_name)+1]
            priceID += 1
            priceIDpos = 10
        
        flag22 = 0
        #struct = struct[:1]
        if flag2 == 1 or flag3 == 1:
            priceID += 1
            priceIDpos = 11
        if map_row_count == target_row:
            x = 0
        level2_nat = df.loc[int(item2[1:])-1][nat_col]
        level2_name = df.loc[int(item2[1:])-1][name_col]
        level2_cant = df.loc[int(item2[1:])-1][cant_col]
        level2_um = df.loc[int(item2[1:])-1][um_col]
        level2_nat = df.loc[int(item2[1:])-1][nat_col]
        flag2 = 1
        if not isinstance(level2_cant,int) and not isinstance(level2_cant,float):
            if level2_cant[0:1] == '=':
                l2cant = level2_cant[1:]
                level2_cant = final_cant(l2cant)
            struct.append([level2_nat,level2_name,level2_cant])
            priceID += 1
            priceIDpos = 12
        
        if map_row_count == target_row:
            print("hola")
        level3_for = final_form_list(item2)
        if isinstance(level3_for[0],int) or isinstance(level3_for[0],float):
            #hemos llegado a los datos finales. Hay que escribir el mapa
            struct.append([level2_nat,level2_name,level3_for[0]])
            final_map.append(map_struct(struct, level2_um, level3_for[1],"c"))
            map_row_count += 1
            print(map_row_count)
            write2 = 1
            #datos finales en este nivel significa que solo hay Capitulo 2 y nada mas, por lo que datos finales son un precio unitario
            continue
        
        #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
        if len(level3_for) == 1:
            itemx= level3_for[0]
            level3_nat = df.loc[int(itemx[1:])-1][nat_col]
            level3_name = df.loc[int(itemx[1:])-1][name_col]
            level3_cant = df.loc[int(itemx[1:])-1][cant_col]
            level3_um = df.loc[int(itemx[1:])-1][um_col]
            level3_pres = df.loc[int(itemx[1:])-1][pres_col]
            level3_for = final_form_list(itemx)
            level3_name_aux = level3_name
            flag22 = 1
            #struct.append([level3_nat,level3_name,level3_cant])
        
            if isinstance(level3_for[0],int) or isinstance(level3_for[0],float):
                #hemos llegado a los datos finales. Hay que escribir el mapa     
                struct.append([level3_nat,level3_name,level3_cant])
                final_map.append(map_struct(struct, level3_um, level3_pres,"d"))
                write2 = 1
                map_row_count += 1
                print(map_row_count)
                #datos finales en este nivel significa que solo hay Capitulo 2 y nada mas, por lo que datos finales son un precio unitario
                if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria":
                    flag1 = 0
                priceID += 1 
                priceIDpos = 13
                continue
            if level3_nat == "Partida":
                flag2 = 1
                if not isinstance(level3_cant,int) and not isinstance(level3_cant,float):
                    if level3_cant[0:1] == '=':
                        l3cant = level3_cant[1:]
                        level3_cant = final_cant(l3cant)
                struct.append([level3_nat,level3_name,level3_cant])
            
        aux3 = 0
        flag1 = 0
        flag3 = 0
        flag33 = 0
        flag4 = 0
        write3 = 0
        write4 = 0
        write5 = 0
        write6 = 0
        write7 = 0
        for item3 in level3_for:
            
            if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
                
                niveles, nombres, cantidades = list(zip(*struct))
                
                if write3 == 1 and (struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida" and struct[-1][2] == 0):
                    priceID += 1
                    priceIDpos = 14

                if write3 == 1 and flag22 == 1:
                    #se creo un nivel adicional
                    struct = struct[:nombres.index(level3_name_aux)+1]
                    if flag3 == 0:
                        priceID += 1
                        priceIDpos = 15
                else:
                    struct = struct[:nombres.index(level2_name)+1]
                    if (flag2== 0 and flag3 == 0) or flag33 == 1:
                        priceID += 1
                        priceIDpos = 16
                    
                    if flag4 == 1 and write4 == 1:
                        priceID += 1
                        priceIDpos = 17
                    
            
            elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and (write3 == 1 or write4 == 1 or write5 == 1) :
                
                niveles, nombres, cantidades = list(zip(*struct))

                struct = struct[:nombres.index(level2_name)+1]
                priceID += 1
                priceIDpos = 18
            
            flag33 = 0
            write3 = 0
            write4 = 0
            write5 = 0
            write6 = 0
            write7 = 0
            if map_row_count == target_row:
                x = 0
            
            level3_nat = df.loc[int(item3[1:])-1][nat_col]
            level3_name = df.loc[int(item3[1:])-1][name_col]
            level3_cant = df.loc[int(item3[1:])-1][cant_col]
            flag3 = 1
            if not isinstance(level3_cant,int) and not isinstance(level3_cant,float):
                if level3_cant[0:1] == '=':
                    l3cant = level3_cant[1:]
                    level3_cant = final_cant(l3cant)
                struct.append([level3_nat,level3_name,level3_cant])
                priceID += 1
                priceIDpos = 19
                
            level3_um = df.loc[int(item3[1:])-1][um_col]
            
            level4_for = final_form_list(item3)
            if isinstance(level4_for[0],int) or isinstance(level4_for[0],float):
                #hemos llegado a los datos finales. Hay que escribir el mapa
                struct.append([level3_nat,level3_name,level4_for[0]])
                final_map.append(map_struct(struct, level3_um, level4_for[1],"e"))
                map_row_count += 1
                print(map_row_count)
                write3 = 1
                #flag2 = 0
                continue
            
            #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
            if len(level4_for) == 1:
                itemx= level4_for[0]
                level4_nat = df.loc[int(itemx[1:])-1][nat_col]
                level4_name = df.loc[int(itemx[1:])-1][name_col]
                level4_cant = df.loc[int(itemx[1:])-1][cant_col]
                level4_um = df.loc[int(itemx[1:])-1][um_col]
                level4_pres = df.loc[int(itemx[1:])-1][pres_col]
                level4_for = final_form_list(itemx)
                level4_name_aux = level4_name
                flag33 = 1
                if isinstance(level4_for[0],int) or isinstance(level4_for[0],float):
                    #hemos llegado a los datos finales. Hay que escribir el mapa
                    struct.append([level4_nat,level4_name,level4_cant])
                    final_map.append(map_struct(struct, level4_um, level4_pres,"f"))
                    write3 = 1
                    map_row_count += 1
                    print(map_row_count)
                    if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria":
                        flag2 = 0
                    continue
                
                if level4_nat == "Partida":
                    flag3 = 1
                    if not isinstance(level4_cant,int) and not isinstance(level4_cant,float):
                        if level4_cant[0:1] == '=':
                            l4cant = level4_cant[1:]
                            level4_cant = final_cant(l4cant)
                    struct.append([level4_nat,level4_name,level4_cant])
            
            flag2 = 0
            aux4 = 0
            flag44 = 0
            flag5 = 0
            #write5 = 0
            #write4 = 0
            for item4 in level4_for:
                
                if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
                    
                    niveles, nombres, cantidades = list(zip(*struct))
                    
                    if write4 == 1 and (struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida" and struct[-1][2] == 0):
                        priceID += 1
                        priceIDpos = 20

                    if write4 == 1 and flag33 == 1:
                        #se creo un nivel adicional

                        #buscamos el último nivel por si hay dos niveles que se llaman iguales
                        #no aplica para cuando el nivel es Recurso, hay que descartar ese item de la lista nombres
                        auxind = 0
                        afirstflag = 0
                        aoffset = 0
                        auxcont = 0
                        for element in nombres:
                            auxcont += 1
                            if niveles[auxcont-1] == "Otros" or niveles[auxcont-1] == "Material" or niveles[auxcont-1] == "Mano de obra" or niveles[auxcont-1] == "Maquinaria":
                                continue
                            if element == level4_name_aux:
                                #auxind muestra el primer índice donde ocurre level4_name_aux (podría llamarse igualq ue level3_name)
                                auxind = nombres.index(element)
                                if afirstflag == 0:
                                    afirstflag = 1
                                    continue
                                aoffset+=1
                        
                        auxind = auxind+aoffset
                        struct = struct[:auxind+1]
                        if flag4 == 0:
                            priceID += 1
                            priceIDpos = 21

                    else:
                        struct = struct[:nombres.index(level3_name)+1]
                        if (flag3== 0 and flag4 == 0) or flag44 == 1:
                            priceID += 1
                            priceIDpos = 22
                        
                        if flag5 == 1 and write5 == 1:
                            priceID += 1
                            priceIDpos = 23

                
                elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and (write4 == 1 or write5 == 1 or write6 == 1 or write7 == 1):
                    
                    niveles, nombres, cantidades = list(zip(*struct))

                    struct = struct[:nombres.index(level3_name)+1]
                    priceID += 1
                    priceIDpos = 24
                
                flag44 = 0
                write4 = 0
                write5 = 0
                write6 = 0
                write7 = 0
                if map_row_count == target_row:
                    x = 0
                
                level4_nat = df.loc[int(item4[1:])-1][nat_col]
                level4_name = df.loc[int(item4[1:])-1][name_col]
                level4_cant = df.loc[int(item4[1:])-1][cant_col]
                flag4 = 1
                if not isinstance(level4_cant,int) and not isinstance(level4_cant,float):
                    if level4_cant[0:1] == '=':
                        l4cant = level4_cant[1:]
                        level4_cant = final_cant(l4cant)
                    struct.append([level4_nat,level4_name,level4_cant])
                    priceID += 1
                    priceIDpos = 25
                level4_um = df.loc[int(item4[1:])-1][um_col]
                
                level5_for = final_form_list(item4)
                
                if isinstance(level5_for[0],int) or isinstance(level5_for[0],float):
                    #hemos llegado a los datos finales. Hay que escribir el mapa
                    struct.append([level4_nat,level4_name,level5_for[0]])
                    final_map.append(map_struct(struct, level4_um, level5_for[1],"g"))
                    write4 = 1
                    map_row_count += 1
                    print(map_row_count)
                    continue
                
                #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
                if len(level5_for) == 1:
                    itemx= level5_for[0]
                    level5_nat = df.loc[int(itemx[1:])-1][nat_col]
                    level5_name = df.loc[int(itemx[1:])-1][name_col]
                    level5_cant = df.loc[int(itemx[1:])-1][cant_col]
                    level5_um = df.loc[int(itemx[1:])-1][um_col]
                    level5_pres = df.loc[int(itemx[1:])-1][pres_col]
                    level5_for = final_form_list(itemx)
                    level5_name_aux = level5_name
                    flag44 = 1
                    if isinstance(level5_for[0],int) or isinstance(level5_for[0],float):
                        #hemos llegado a los datos finales. Hay que escribir el mapa
                        struct.append([level5_nat,level5_name,level5_cant])
                        final_map.append(map_struct(struct, level5_um, level5_pres,"h"))
                        map_row_count += 1
                        if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria":
                            flag3 = 0
                        continue
                    
                    if level5_nat == "Partida":
                        flag4 = 1
                        if not isinstance(level5_cant,int) and not isinstance(level5_cant,float):
                            if level5_cant[0:1] == '=':
                                l5cant = level5_cant[1:]
                                level5_cant = final_cant(l5cant)
                        struct.append([level5_nat,level5_name,level5_cant])
                
                
                flag3 = 0
                flag55 = 0
                aux5 = 0
                write4 = 0
                for item5 in level5_for:
                    
                    if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
                        
                        niveles, nombres, cantidades = list(zip(*struct))
                        
                        if write5 == 1 and (struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida" and struct[-1][2] == 0):
                            priceID += 1
                            priceIDpos = 26

                        if write5 == 1 and flag44 == 1:
                            #se creo un nivel adicional
                            struct = struct[:nombres.index(level5_name_aux)+1]
                            
                            if flag5 == 0:
                                priceID += 1
                                priceIDpos = 27
                        else:
                            struct = struct[:nombres.index(level4_name)+1]
                            if flag55 == 1 or write6 == 1 or write7 == 1:
                                priceID += 1
                                priceIDpos = 28  
                            
                    
                    elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and (write5 == 1 or write6 == 1 or write7 == 1) :
                        
                        niveles, nombres, cantidades = list(zip(*struct))

                        struct = struct[:nombres.index(level4_name)+1]
                        priceID += 1
                        priceIDpos = 29
                    
                    flag55 = 0  
                    write5 = 0
                    write6 = 0
                    write7 = 0
                    write8 = 0
                    write9 = 0
                    if map_row_count == target_row:
                        x = 0
                     
                    level5_nat = df.loc[int(item5[1:])-1][nat_col]
                    level5_name = df.loc[int(item5[1:])-1][name_col]
                    level5_cant = df.loc[int(item5[1:])-1][cant_col]
                    flag5 = 1
                    if not isinstance(level5_cant,int) and not isinstance(level5_cant,float):
                        if level5_cant[0:1] == '=':
                            l5cant = level5_cant[1:]
                            level5_cant = final_cant(l5cant)
                        struct.append([level5_nat,level5_name,level5_cant])
                        priceID += 1
                        priceIDpos = 30
                    level5_um = df.loc[int(item5[1:])-1][um_col]
                    
                    level6_for = final_form_list(item5)
                    
                    if isinstance(level6_for[0],int) or isinstance(level6_for[0],float):
                        #hemos llegado a los datos finales. Hay que escribir el mapa. Recurso es level 5. Level 4 es partida 1, se agregan dos partidas mas
                        struct.append([level5_nat,level5_name,level6_for[0]])
                        final_map.append(map_struct(struct, level5_um, level6_for[1],"i"))
                        write5 = 1
                        map_row_count += 1
                        print(map_row_count)
                        continue
                    
                    #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
                    if len(level6_for) == 1:
                        itemx= level6_for[0]
                        level6_nat = df.loc[int(itemx[1:])-1][nat_col]
                        level6_name = df.loc[int(itemx[1:])-1][name_col]
                        level6_cant = df.loc[int(itemx[1:])-1][cant_col]
                        level6_um = df.loc[int(itemx[1:])-1][um_col]
                        level6_pres = df.loc[int(itemx[1:])-1][pres_col]
                        level6_for = final_form_list(itemx)
                        level6_name_aux = level6_name
                        flag55 = 1
                        if isinstance(level6_for[0],int) or isinstance(level6_for[0],float):
                            #hemos llegado a los datos finales. Hay que escribir el mapa
                            struct.append([level6_nat,level6_name,level6_cant])
                            final_map.append(map_struct(struct, level6_um, level6_pres,"j"))
                            map_row_count += 1
                            if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria":
                                flag4 = 0
                            continue
                        
                        if level6_nat == "Partida":
                            flag5 = 1
                            if not isinstance(level6_cant,int) and not isinstance(level6_cant,float):
                                if level6_cant[0:1] == '=':
                                    l6cant = level6_cant[1:]
                                    level6_cant = final_cant(l6cant)
                            struct.append([level6_nat,level6_name,level6_cant])
                        
                    aux6 = 0
                    flag4 = 0
                    flag6 = 0
                    flag7 = 0
                    flag66 = 0
                    flag77 = 0
                    write5 = 0
                    for item6 in level6_for:
                        
                        if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
                            
                            niveles, nombres, cantidades = list(zip(*struct))
                            
                            if write6 == 1 and (struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida" and struct[-1][2] == 0):
                                priceID += 1
                                priceIDpos = 31

                            if write6 == 1 and flag55 == 1:
                                #se creo un nivel adicional
                                priceID += 1
                                priceIDpos = 32
                                struct = struct[:nombres.index(level6_name_aux)+1]
                            else:
                                struct = struct[:nombres.index(level5_name)+1]
                                if flag66 == 1 or write7 == 1 or write8 == 1:
                                    priceID += 1
                                    priceIDpos = 33
                        
                        elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and (write6 == 1 or write7 == 1) :
                            
                            niveles, nombres, cantidades = list(zip(*struct))

                            struct = struct[:nombres.index(level5_name)+1]
                            priceID += 1
                            priceIDpos = 34

                        flag66 = 0
                        write6 = 0
                        write7 = 0
                        if map_row_count == target_row:
                            x = 0
                            
                        level6_nat = df.loc[int(item6[1:])-1][nat_col]
                        level6_name = df.loc[int(item6[1:])-1][name_col]
                        level6_cant = df.loc[int(item6[1:])-1][cant_col]
                        if not isinstance(level6_cant,int) and not isinstance(level6_cant,float):
                            if level6_cant[0:1] == '=':
                                l6cant = level6_cant[1:]
                                level6_cant = final_cant(l6cant)
                            struct.append([level6_nat,level6_name,level6_cant])
                            priceID += 1
                            priceIDpos = 35
                        level6_um = df.loc[int(item6[1:])-1][um_col]
                        
                        level7_for = final_form_list(item6)
                        
                        if isinstance(level7_for[0],int) or isinstance(level7_for[0],float):
                            #hemos llegado a los datos finales. Hay que escribir el mapa
                            struct.append([level6_nat,level6_name,level7_for[0]])
                            final_map.append(map_struct(struct, level6_um, level7_for[1],"k"))
                            write6 = 1
                            map_row_count += 1
                            print(map_row_count)
                            continue
                        
                        #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
                        if len(level7_for) == 1:
                            itemx= level7_for[0]
                            level7_nat = df.loc[int(itemx[1:])-1][nat_col]
                            level7_name = df.loc[int(itemx[1:])-1][name_col]
                            level7_cant = df.loc[int(itemx[1:])-1][cant_col]
                            level7_um = df.loc[int(itemx[1:])-1][um_col]
                            level7_pres = df.loc[int(itemx[1:])-1][pres_col]
                            level7_for = final_form_list(itemx)
                            level7_name_aux = level7_name
                            flag66 = 1
                            
                            if isinstance(level7_for[0],int) or isinstance(level7_for[0],float):
                                #hemos llegado a los datos finales. Hay que escribir el mapa
                                struct.append([level7_nat,level7_name,level7_cant])
                                final_map.append(map_struct(struct, level7_um, level7_pres,"l"))
                                map_row_count += 1
                                print(map_row_count)
                                if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria":
                                    flag5 = 0
                                continue
                            
                            if level7_nat == "Partida":
                                flag6 = 1
                                if not isinstance(level7_cant,int) and not isinstance(level7_cant,float):
                                    if level7_cant[0:1] == '=':
                                        l7cant = level7_cant[1:]
                                        level7_cant = final_cant(l7cant)
                                struct.append([level7_nat,level7_name,level7_cant])
                        
                        write6 = 0
                        flag5 = 0
                        for item7 in level7_for:
                            if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
                                
                                niveles, nombres, cantidades = list(zip(*struct))
                                
                                if write7 == 1 and (struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida" and struct[-1][2] == 0):
                                    priceID += 1
                                    priceIDpos = 36

                                if write7 == 1 and flag66 == 1:
                                    #se creo un nivel adicional
                                    struct = struct[:nombres.index(level7_name_aux)+1]
                                    if flag6 == 0:
                                        priceID += 1
                                        priceIDpos = 37
                                    
                                else:
                                    struct = struct[:nombres.index(level6_name)+1]
                                    if flag77 == 1 or write8 == 1:
                                        priceID += 1
                                        priceIDpos = 38
                            
                            elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and write7 == 1 :
                                
                                niveles, nombres, cantidades = list(zip(*struct))

                                struct = struct[:nombres.index(level6_name)+1]
                                priceID += 1
                                priceIDpos = 39
                            
                            flag77 = 0
                            write7 = 0
                            write8 = 0
                            if map_row_count == target_row:
                                x = 0
                                
                            level7_nat = df.loc[int(item7[1:])-1][nat_col]
                            level7_name = df.loc[int(item7[1:])-1][name_col]
                            level7_cant = df.loc[int(item7[1:])-1][cant_col]
                            if not isinstance(level7_cant,int) and not isinstance(level7_cant,float):
                                if level7_cant[0:1] == '=':
                                     l7cant = level7_cant[1:]
                                     level7_cant = final_cant(l7cant)
                                struct.append([level7_nat,level7_name,level7_cant])   
                                priceID += 1
                                priceIDpos = 40                            
                            level7_um = df.loc[int(item7[1:])-1][um_col]
                            
                            level8_for = final_form_list(item7)
                            
                            if isinstance(level8_for[0],int) or isinstance(level8_for[0],float):
                                #hemos llegado a los datos finales. Hay que escribir el mapa
                                struct.append([level7_nat,level7_name,level8_for[0]])
                                final_map.append(map_struct(struct, level7_um, level8_for[1],"m"))
                                write7 = 1
                                map_row_count += 1
                                print(map_row_count)
                                continue
                            
                            #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
                            if len(level8_for) == 1:
                                itemx= level8_for[0]
                                level8_nat = df.loc[int(itemx[1:])-1][nat_col]
                                level8_name = df.loc[int(itemx[1:])-1][name_col]
                                level8_cant = df.loc[int(itemx[1:])-1][cant_col]
                                level8_um = df.loc[int(itemx[1:])-1][um_col]
                                level8_pres = df.loc[int(itemx[1:])-1][pres_col]
                                level8_for = final_form_list(itemx)
                                level8_name_aux = level8_name
                                flag77 = 1
                            
                                if isinstance(level8_for[0],int) or isinstance(level8_for[0],float):
                                    #hemos llegado a los datos finales. Hay que escribir el mapa
                                    struct.append([level8_nat,level8_name,level8_cant])
                                    final_map.append(map_struct(struct, level8_um, level8_pres,"n"))
                                    map_row_count += 1
                                    print(map_row_count)
                                    continue
                                
                            flag6 = 0
                            write7 = 0
                                    
                            for item8 in level8_for:
                                if struct[-1][0] == "Otros" or struct[-1][0] == "Material" or struct[-1][0] == "Mano de obra" or struct[-1][0] == "Maquinaria" or ((struct[-1][0] == "Capítulo" or struct[-1][0] == "Partida") and struct[-1][2] == 0):
                                    
                                    niveles, nombres, cantidades = list(zip(*struct))

                                    if write8 == 1 and flag77 == 1:
                                        #se creo un nivel adicional
                                        struct = struct[:nombres.index(level8_name_aux)+1]
                                        if flag7 == 0:
                                            priceID += 1
                                            priceIDpos = 41
                                    else:
                                        struct = struct[:nombres.index(level7_name)+1]
                                        if write9 == 1:
                                            priceID += 1
                                            priceIDpos = 42
                                
                                elif (struct[-1][0] == "Partida" or struct[-1][0] == "Capítulo") and write8 == 1 :
                                    
                                    niveles, nombres, cantidades = list(zip(*struct))

                                    struct = struct[:nombres.index(level7_name)+1]
                                    priceID += 1
                                    priceIDpos = 43
                                
                                
                                write8 = 0
                                write9 = 0
                                if map_row_count == target_row:
                                    x = 0
                                    
                                level8_nat = df.loc[int(item8[1:])-1][nat_col]
                                level8_name = df.loc[int(item8[1:])-1][name_col]
                                level8_cant = df.loc[int(item8[1:])-1][cant_col]
                                if not isinstance(level8_cant,int) and not isinstance(level8_cant,float):
                                    if level8_cant[0:1] == '=':
                                        l8cant = level8_cant[1:]
                                        level8_cant = final_cant(l8cant)
                                    struct.append([level8_nat,level8_name,level8_cant]) 
                                    priceID += 1
                                    priceIDpos = 44
                                level8_um = df.loc[int(item8[1:])-1][um_col]
                                
                                level9_for = final_form_list(item8)
                                
                                if isinstance(level9_for[0],int) or isinstance(level9_for[0],float):
                                    #hemos llegado a los datos finales. Hay que escribir el mapa
                                    struct.append([level8_nat,level8_name,level9_for[0]])
                                    final_map.append(map_struct(struct, level8_um, level9_for[1],"o"))
                                    write8 = 1
                                    map_row_count += 1
                                    print(map_row_count)
                                    continue
                                
                                #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
                                if len(level9_for) == 1:
                                    itemx= level9_for[0]
                                    level9_name = df.loc[int(itemx[1:])-1][name_col]
                                    level9_cant = df.loc[int(itemx[1:])-1][cant_col]
                                    level9_um = df.loc[int(itemx[1:])-1][um_col]
                                    level9_pres = df.loc[int(itemx[1:])-1][pres_col]
                                    level9_for = final_form_list(itemx)
                                
                                    if isinstance(level9_for[0],int) or isinstance(level9_for[0],float):
                                        #hemos llegado a los datos finales. Hay que escribir el mapa
                                        total_cant8 = level1_cant*level2_cant*level3_cant*level4_cant*level5_cant*level6_cant*level7_cant*level8_cant*level9_cant
                                        total_pres8 = total_cant8*level9_pres
                                        final_map.append((level1_name,level2_name,level3_name,level4_name,level5_name,level6_name, level7_name,level8_name, level9_name, level9_um,level1_cant,level2_cant,level3_cant,level4_cant,level5_cant,level6_cant,level7_cant,level8_cant, level9_cant,level9_pres,total_cant8, total_pres8,"p"))
                                        map_row_count += 1
                                        print(map_row_count)
                                        continue
                                
                                write8 = 0
                                for item9 in level9_for:
                                    struct = struct[:8] 
                                    if map_row_count == target_row:
                                        x = 0
                                        
                                    level9_nat = df.loc[int(item9[1:])-1][nat_col]
                                    level9_name = df.loc[int(item9[1:])-1][name_col]
                                    level9_cant = df.loc[int(item9[1:])-1][cant_col]
                                    if not isinstance(level9_cant,int) and not isinstance(level9_cant,float):
                                        if level9_cant[0:1] == '=':
                                            l9cant = level9_cant[1:]
                                            level9_cant = final_cant(l9cant)
                                        struct.append([level9_nat,level9_name,level9_cant]) 
                                        priceID += 1
                                        priceIDpos = 45
                                    level9_um = df.loc[int(item9[1:])-1][um_col]
                                    
                                    level10_for = final_form_list(item9)
                                    
                                    if isinstance(level10_for[0],int) or isinstance(level10_for[0],float):
                                        #hemos llegado a los datos finales. Hay que escribir el mapa
                                        struct.append([level9_nat,level9_name,level10_for[0]])
                                        final_map.append(map_struct(struct, level9_um, level10_for[1],"q"))
                                        write9 = 1
                                        map_row_count += 1
                                        print(map_row_count)
                                        continue
                                    
                                    #revisamos el largo de la lista. Si tiene sólo un elemento, pueden ser datos finales
                                    if len(level10_for) == 1:
                                        itemx= level10_for[0]
                                        level10_name = df.loc[int(itemx[1:])-1][name_col]
                                        level10_cant = df.loc[int(itemx[1:])-1][cant_col]
                                        level10_um = df.loc[int(itemx[1:])-1][um_col]
                                        level10_pres = df.loc[int(itemx[1:])-1][pres_col]
                                        level10_for = final_form_list(itemx)
                                    
                                        if isinstance(level10_for[0],int) or isinstance(level10_for[0],float):
                                            #hemos llegado a los datos finales. Hay que escribir el mapa
                                            total_cant8 = level1_cant*level2_cant*level3_cant*level4_cant*level5_cant*level6_cant*level7_cant*level8_cant*level9_cant
                                            total_pres8 = total_cant8*level9_pres
                                            final_map.append((level1_name,level2_name,level3_name,level4_name,level5_name,level6_name, level7_name,level8_name, level9_name, level9_um,level1_cant,level2_cant,level3_cant,level4_cant,level5_cant,level6_cant,level7_cant,level8_cant, level9_cant,level9_pres,total_cant8, total_pres8,"r"))
                                            map_row_count += 1
                                            print(map_row_count)
                                            continue
                                     
                                        write9 = 0

pd.DataFrame(final_map).to_excel(path+"/resultado_priceID2.xlsx", index =False, header = False)


