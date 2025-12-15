# -*- coding: utf-8 -*-
"""
Created on Mon Dec 15 16:26:40 2025

@author: jfcog
"""

"""Generalized parser for Presto Excel exports.

The original ``PrestoV9 Price ID.py`` script depended on hand-tuned
loops and counters to follow la estructura de capítulos, partidas y
recursos. Este módulo crea un grafo de dependencias a partir de las
fórmulas del Excel, recorre recursivamente las ramas desde la celda de
presupuesto total y emite el mismo esquema de salida (hasta 6 capítulos,
4 partidas y los totales por recurso) sin fijar la profundidad del
archivo de entrada.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from functools import reduce
from operator import mul
from typing import Dict, List, Optional, Sequence, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

# Column positions in the Presto export (0-based, aligned with the
# original script defaults)
NATURE_COL = 1
UNIT_COL = 2
NAME_COL = 3
QUANTITY_COL = 4
PRICE_COL = 5
HEADER_ROW_INDEX = 2  # Row that contains the column titles

RESOURCE_TYPES = {"Otros", "Material", "Mano de obra", "Maquinaria"}
COLUMN_REF_RE = re.compile(r"[A-Z]+\d+")
RANGE_RE = re.compile(r"([A-Z]+\d+):([A-Z]+\d+)")


@dataclass
class RowMetadata:
    row_index: int  # zero-based row index in the DataFrame
    nature: str
    name: str
    unit: str
    quantity_cell: str
    price_cell: str


class DependencyGraph:
    """Tracks which cells depend on others inside the sheet."""

    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.graph: Dict[str, Set[str]] = {}
        self._build_graph()

    def _coord(self, row_idx: int, col_idx: int) -> str:
        return f"{get_column_letter(col_idx + 1)}{row_idx + 1}"

    def _expand_range(self, start: str, end: str) -> List[str]:
        min_col, min_row, max_col, max_row = range_boundaries(f"{start}:{end}")
        coords = []
        for col in range(min_col, max_col + 1):
            for row in range(min_row, max_row + 1):
                coords.append(f"{get_column_letter(col)}{row}")
        return coords

    def _extract_references(self, formula: str) -> List[str]:
        refs: List[str] = []
        for start, end in RANGE_RE.findall(formula):
            refs.extend(self._expand_range(start, end))
        formula_without_ranges = RANGE_RE.sub("", formula)
        refs.extend(COLUMN_REF_RE.findall(formula_without_ranges))
        return refs

    def _build_graph(self) -> None:
        for row_idx, row in self.df.iterrows():
            for col_idx, value in row.items():
                if isinstance(value, str) and (value.startswith("=") or "SUM" in value or "ROUND" in value):
                    coord = self._coord(row_idx, col_idx)
                    refs = set(self._extract_references(value))
                    self.graph[coord] = refs

    def children(self, coord: str) -> Set[str]:
        return self.graph.get(coord, set())


class FormulaEvaluator:
    """Evaluates Excel-like formulas using the dependency graph."""

    def __init__(self, df: pd.DataFrame, graph: DependencyGraph):
        self.df = df
        self.graph = graph
        self.cache: Dict[str, float] = {}

    def _coord_value(self, coord: str) -> Optional[float]:
        column_letter = re.findall(r"[A-Z]+", coord)[0]
        row_number = int(re.findall(r"\d+", coord)[0])
        col_idx = column_index_from_string(column_letter) - 1
        row_idx = row_number - 1
        value = self.df.loc[row_idx][col_idx]
        if isinstance(value, (int, float)):
            return float(value)
        return None

    def evaluate(self, coord: str) -> float:
        if coord in self.cache:
            return self.cache[coord]
        value = self._coord_value(coord)
        if value is not None:
            self.cache[coord] = value
            return value

        column_letter = re.findall(r"[A-Z]+", coord)[0]
        row_number = int(re.findall(r"\d+", coord)[0]) - 1
        col_idx = column_index_from_string(column_letter) - 1
        raw_formula = self.df.loc[row_number][col_idx]
        if not isinstance(raw_formula, str):
            raise ValueError(f"Cannot evaluate cell {coord}: {raw_formula}")
        expression = raw_formula[1:] if raw_formula.startswith("=") else raw_formula

        def replace_range(match: re.Match[str]) -> str:
            start, end = match.groups()
            coords = self.graph._expand_range(start, end)
            return str(sum(self.evaluate(c) for c in coords))

        expression = RANGE_RE.sub(replace_range, expression)

        def replace_coord(match: re.Match[str]) -> str:
            return str(self.evaluate(match.group(0)))

        expression = COLUMN_REF_RE.sub(replace_coord, expression)
        expression = expression.replace("^", "**")
        value = float(eval(expression, {"__builtins__": {}, "SUM": sum, "ROUND": round}))
        self.cache[coord] = value
        return value


class TraversalContext:
    def __init__(self, chapters: List[RowMetadata], partidas: List[RowMetadata]):
        self.chapters = chapters
        self.partidas = partidas

    def add(self, meta: RowMetadata) -> "TraversalContext":
        if meta.nature == "Capítulo":
            return TraversalContext(self.chapters + [meta], self.partidas)
        if meta.nature == "Partida":
            return TraversalContext(self.chapters, self.partidas + [meta])
        return self


def _coord(row_idx: int, col_idx: int) -> str:
    return f"{get_column_letter(col_idx + 1)}{row_idx + 1}"


def build_row_metadata(df: pd.DataFrame) -> Dict[int, RowMetadata]:
    metadata: Dict[int, RowMetadata] = {}
    for idx, row in df.iterrows():
        nature = row.get(NATURE_COL)
        name = row.get(NAME_COL)
        unit = row.get(UNIT_COL)
        if not isinstance(nature, str) or not isinstance(name, str):
            continue
        quantity_cell = _coord(idx, QUANTITY_COL)
        price_cell = _coord(idx, PRICE_COL)
        metadata[idx] = RowMetadata(idx, nature, name, str(unit or ""), quantity_cell, price_cell)
    return metadata


def _product(values: Sequence[float]) -> float:
    return reduce(mul, values, 1.0)


def _pad(values: Sequence[str], length: int) -> List[str]:
    values = list(values)
    while len(values) < length:
        values.append("")
    return values[:length]


def find_budget_cell(df: pd.DataFrame) -> str:
    pres_col = None
    for idx, value in enumerate(df.loc[HEADER_ROW_INDEX]):
        if value == "Pres":
            pres_col = idx
            break
    if pres_col is None:
        raise ValueError("No se encontró la columna 'Pres'")
    for row_idx in range(len(df) - 1, -1, -1):
        cell_value = df.loc[row_idx][pres_col]
        if cell_value not in (None, ""):
            return _coord(row_idx, pres_col)
    raise ValueError("No se encontró la celda de presupuesto total")


def emit_row(
    context: TraversalContext,
    resource_meta: RowMetadata,
    evaluator: FormulaEvaluator,
    editor_tag: str,
) -> List[object]:
    chapter_names = [meta.name for meta in context.chapters]
    partida_names = [meta.name for meta in context.partidas]
    chapter_qty = [evaluator.evaluate(meta.quantity_cell) for meta in context.chapters]
    partida_qty = [evaluator.evaluate(meta.quantity_cell) for meta in context.partidas]

    resource_qty = evaluator.evaluate(resource_meta.quantity_cell)
    resource_price = evaluator.evaluate(resource_meta.price_cell)

    total_quantity = _product(chapter_qty + partida_qty + [resource_qty])
    total_cost = total_quantity * resource_price

    row = []
    row.extend(_pad(chapter_names, 6))
    row.extend(_pad(partida_names, 4))
    row.append(resource_meta.name)
    row.append(resource_meta.nature)
    row.append(resource_meta.unit)
    row.extend(_pad([str(q) for q in chapter_qty], 6))
    row.extend(_pad([str(q) for q in partida_qty], 4))
    row.append(resource_qty)
    row.append(resource_price)
    row.append(total_quantity)
    row.append(total_cost)
    row.append(editor_tag)
    return row


def traverse_resources(
    df: pd.DataFrame,
    graph: DependencyGraph,
    evaluator: FormulaEvaluator,
    metadata: Dict[int, RowMetadata],
    start_cell: str,
    editor_tag: str = "auto",
) -> List[List[object]]:
    visited: Set[str] = set()
    output: List[List[object]] = []

    def _walk(cell_ref: str, context: TraversalContext) -> None:
        if cell_ref in visited:
            return
        visited.add(cell_ref)

        # Identify row metadata for this cell, if any
        row_number = int(re.findall(r"\d+", cell_ref)[0]) - 1
        meta = metadata.get(row_number)
        next_context = context.add(meta) if meta else context

        children = graph.children(cell_ref)
        if children:
            for child in children:
                _walk(child, next_context)
            return

        if meta and meta.nature in RESOURCE_TYPES:
            output.append(emit_row(next_context, meta, evaluator, editor_tag))

    _walk(start_cell, TraversalContext([], []))
    return output


def generate_resource_map(workbook_path: str, sheet_name: Optional[str] = None) -> List[List[object]]:
    wb = load_workbook(workbook_path, data_only=False)
    sheet = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    df = pd.DataFrame(sheet.values)

    metadata = build_row_metadata(df)
    graph = DependencyGraph(df)
    evaluator = FormulaEvaluator(df, graph)
    budget_cell = find_budget_cell(df)

    header = (
        "Capitulo 1",
        "Capitulo 2 ",
        "Capitulo 3",
        "Capitulo 4",
        "Capitulo 5",
        "Capitulo 6",
        "Partida 1",
        "Partida 2",
        "Partida 3",
        "Partida 4",
        "Recurso",
        "Tipo Recurso",
        "UM Recurso",
        "Cantidad C1",
        "Cantidad C2",
        "Cantidad C3",
        "Cantidad C4",
        "Cantidad C5",
        "Cantidad C6",
        "Cantidad P1",
        "Cantidad P2",
        "Cantidad P3",
        "Cantidad P4",
        "Cantidad Recurso",
        "Precio Recurso",
        "Total Cant",
        "Presupuesto Total",
        "Editor",
    )

    rows = [list(header)]
    rows.extend(traverse_resources(df, graph, evaluator, metadata, budget_cell))
    return rows


def export_to_csv(workbook_path: str, output_csv: str, sheet_name: Optional[str] = None) -> None:
    rows = generate_resource_map(workbook_path, sheet_name=sheet_name)
    export_df = pd.DataFrame(rows)
    export_df.to_csv(output_csv, index=False, header=False)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Reconstruye totales por recurso desde un Excel de Presto.")
    parser.add_argument("excel_path", help="Ruta al archivo Excel exportado desde Presto")
    parser.add_argument(
        "--sheet",
        dest="sheet_name",
        default=None,
        help="Nombre de la pestaña a procesar (por defecto la primera)",
    )
    parser.add_argument(
        "--output",
        dest="output_csv",
        default="resource_totals.csv",
        help="Ruta del CSV de salida con el mismo formato que el script original",
    )

    args = parser.parse_args()
    export_to_csv(args.excel_path, args.output_csv, sheet_name=args.sheet_name)
